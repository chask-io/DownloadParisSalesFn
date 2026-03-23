"""
Sandbox: B2B Paris — Complete download flow.
After DESCARGAR, a "Descargar Archivo" dialog appears with a download link.
Click the link to get the file.
"""

import json
import time
from datetime import datetime, timedelta
from pathlib import Path

from playwright.sync_api import sync_playwright
from chask_sdk import ChaskClient

profiles = json.loads(Path.home().joinpath(".chask/profiles.json").read_text())
SDK_TOKEN = profiles["profiles"]["prod-top"]["apiKey"]
client = ChaskClient(token=SDK_TOKEN)

B2B_USER = client.get_secret("0a6ed1f7-2c87-4c3a-946f-69df1c1955ce").reveal()
B2B_PASS = client.get_secret("6528a4b3-04cc-4d26-8fa8-00ec8cc55149").reveal()

B2B_PORTAL = "https://www.cenconlineb2b.com/ParisCL/BBRe-commerce/main"

SNAPSHOTS_DIR = Path(__file__).parent / "snapshots"
SNAPSHOTS_DIR.mkdir(exist_ok=True)
DOWNLOADS_DIR = Path(__file__).parent / "downloads"
DOWNLOADS_DIR.mkdir(exist_ok=True)


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")


def snapshot(page, name):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    page.screenshot(path=str(SNAPSHOTS_DIR / f"{ts}_{name}.png"), full_page=True)
    log(f"  Screenshot: {ts}_{name}.png")


def login(page):
    page.goto(B2B_PORTAL, wait_until="domcontentloaded", timeout=30000)
    time.sleep(2)
    page.fill("#username", B2B_USER, timeout=10000)
    page.fill("#password", B2B_PASS)
    page.click("#kc-login")
    for _ in range(20):
        time.sleep(3)
        if "paris" in page.title().lower() and "oops" not in page.title().lower():
            return True
    return False


def wait_vaadin(page, timeout=60):
    for _ in range(timeout // 3):
        time.sleep(3)
        count = page.evaluate("() => document.querySelectorAll('[class*=\"v-\"]').length")
        if count > 50:
            log(f"Vaadin ready ({count} elements)")
            return True
    return False


def click_vaadin_menu(page, text):
    return page.evaluate(f"""() => {{
        const captions = document.querySelectorAll('.v-menubar-menuitem-caption');
        for (const cap of captions) {{
            if (cap.textContent.trim() === '{text}') {{
                cap.closest('.v-menubar-menuitem').click();
                return true;
            }}
        }}
        return false;
    }}""")


def click_submenu(page, text):
    return page.evaluate(f"""() => {{
        const popup = document.querySelector('.v-menubar-popup');
        if (!popup) return false;
        const items = popup.querySelectorAll('.v-menubar-menuitem-caption');
        for (const item of items) {{
            if (item.textContent.trim() === '{text}') {{
                item.closest('.v-menubar-menuitem').click();
                return true;
            }}
        }}
        return false;
    }}""")


def main():
    log("Starting B2B Paris — full download flow...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            accept_downloads=True
        )
        page = context.new_page()

        downloads = []
        page.on("download", lambda d: (downloads.append(d), log(f"  DOWNLOAD: {d.suggested_filename}")))

        try:
            if not login(page):
                return
            if not wait_vaadin(page):
                return
            time.sleep(2)

            # Navigate
            log("Comercial > Informe de Ventas...")
            click_vaadin_menu(page, "Comercial")
            time.sleep(2)
            click_submenu(page, "Informe de Ventas")
            time.sleep(5)

            # Generate report
            log("Generating report...")
            page.evaluate("""() => {
                const buttons = document.querySelectorAll('.v-button');
                for (const btn of buttons) {
                    if (btn.textContent.trim().includes('Generar Informe')) { btn.click(); return; }
                }
            }""")
            for _ in range(20):
                time.sleep(3)
                if page.evaluate("() => document.querySelectorAll('.v-grid-body .v-grid-row').length") > 0:
                    log("Report loaded")
                    break

            # Open download popup
            log("Download popup...")
            page.evaluate("""() => {
                const buttons = document.querySelectorAll('.v-button.toolbar-button');
                for (const btn of buttons) {
                    const img = btn.querySelector('img.v-icon');
                    if (img && img.src.includes('DownloadPrimary')) { btn.click(); return; }
                }
            }""")
            time.sleep(2)

            # Click "Descarga Venta Histórica a Nivel de Producto-Local"
            log("Click Venta Histórica option...")
            page.evaluate("""() => {
                const popup = document.querySelector('.v-popupview-popup .popupContent');
                if (!popup) return;
                const buttons = popup.querySelectorAll('.v-button');
                for (const btn of buttons) {
                    if (btn.textContent.trim().includes('Producto-Local')) { btn.click(); return; }
                }
            }""")
            time.sleep(3)

            # ── Modal appears: "Descarga de Datos Fuente de Ventas" ──
            log("\nModal appeared, exploring...")
            snapshot(page, "01_modal")

            # Select the first radio (custom date range) — it should already be selected
            # Set dates in modal to last 7 days
            today = datetime.now()
            desde = (today - timedelta(days=7)).strftime("%d-%m-%Y")
            hasta = today.strftime("%d-%m-%Y")

            log(f"Setting modal dates: {desde} to {hasta}")
            page.evaluate(f"""() => {{
                const win = document.querySelector('.v-window');
                if (!win) return;
                // Select first radio button (custom date range)
                const radios = win.querySelectorAll('input[type="radio"]');
                if (radios[0]) radios[0].click();
            }}""")
            time.sleep(1)

            # Set dates via JS
            page.evaluate(f"""() => {{
                const win = document.querySelector('.v-window');
                const dateInputs = win.querySelectorAll('.v-textfield.v-datefield-textfield');
                // First two are the Desde/Hasta for custom range
                for (let i = 0; i < 2; i++) {{
                    const input = dateInputs[i];
                    input.removeAttribute('disabled');
                    input.value = i === 0 ? '{desde}' : '{hasta}';
                    input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    input.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                }}
            }}""")
            time.sleep(1)
            snapshot(page, "02_dates_set")

            # Click DESCARGAR using Playwright locator (not JS)
            log("Clicking DESCARGAR...")
            descargar_btn = page.locator(".v-window .v-button:has-text('Descargar')").first
            descargar_btn.click(force=True)

            # Wait for "Descargar Archivo" dialog with download link
            log("Waiting for file generation...")
            for i in range(90):
                time.sleep(2)

                # Check for the "Descargar Archivo" window with link
                link_info = page.evaluate("""() => {
                    const windows = document.querySelectorAll('.v-window');
                    for (const win of windows) {
                        const header = win.querySelector('.v-window-header');
                        if (header && header.textContent.trim().includes('Descargar Archivo')) {
                            const links = win.querySelectorAll('a');
                            const labels = win.querySelectorAll('.v-label');
                            return {
                                found: true,
                                header: header.textContent.trim(),
                                links: Array.from(links).map(a => ({text: a.textContent.trim(), href: a.href})),
                                labels: Array.from(labels).map(l => l.textContent.trim()),
                                html: win.innerHTML.substring(0, 500)
                            };
                        }
                    }
                    return { found: false };
                }""")

                if link_info.get("found"):
                    log(f"\nFile ready dialog found!")
                    log(f"  Header: {link_info['header']}")
                    log(f"  Labels: {link_info['labels']}")
                    log(f"  Links: {link_info['links']}")
                    snapshot(page, "03_file_ready")

                    # Click the download link
                    if link_info['links']:
                        link_href = link_info['links'][0]['href']
                        log(f"  Clicking download link: {link_href[:80]}")

                        with page.expect_download(timeout=60000) as download_info:
                            page.evaluate(f"""() => {{
                                const windows = document.querySelectorAll('.v-window');
                                for (const win of windows) {{
                                    const header = win.querySelector('.v-window-header');
                                    if (header && header.textContent.trim().includes('Descargar Archivo')) {{
                                        const link = win.querySelector('a');
                                        if (link) {{ link.click(); return true; }}
                                    }}
                                }}
                                return false;
                            }}""")

                        dl = download_info.value
                        log(f"\nFile downloaded: {dl.suggested_filename}")
                        save_path = DOWNLOADS_DIR / dl.suggested_filename
                        dl.save_as(str(save_path))
                        log(f"Saved: {save_path} ({save_path.stat().st_size} bytes)")

                        # Analyze
                        log("\n--- FILE ANALYSIS ---")
                        if save_path.suffix in ('.xlsx', '.xls'):
                            import openpyxl
                            wb = openpyxl.load_workbook(str(save_path), data_only=True)
                            for name in wb.sheetnames:
                                ws = wb[name]
                                log(f"Sheet: {name}, rows={ws.max_row}, cols={ws.max_column}")
                                log("Headers:")
                                for cell in ws[1]:
                                    if cell.value:
                                        log(f"  {cell.column_letter}: {cell.value}")
                                log("Sample rows:")
                                for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
                                    log(f"  {list(row)[:12]}")
                        elif save_path.suffix == '.csv':
                            lines = save_path.read_text(errors='replace').split('\n')[:6]
                            for line in lines:
                                log(f"  {line[:200]}")
                        else:
                            log(f"Format: {save_path.suffix}, size: {save_path.stat().st_size}")
                    break

                # Check for other downloads
                if downloads:
                    dl = downloads[0]
                    log(f"Browser download: {dl.suggested_filename}")
                    save_path = DOWNLOADS_DIR / dl.suggested_filename
                    dl.save_as(str(save_path))
                    log(f"Saved: {save_path}")
                    break

                if i % 15 == 0:
                    log(f"  [{i*2}s] Still generating...")
            else:
                log("Timeout waiting for file")
                snapshot(page, "04_timeout")

            log("\nDONE")

        except Exception as e:
            log(f"ERROR: {e}")
            snapshot(page, "error")
            raise
        finally:
            browser.close()


if __name__ == "__main__":
    main()
